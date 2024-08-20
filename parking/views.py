# views.py
import datetime
import io
import logging
from django.db.models import Sum, Count
from django.utils.dateparse import parse_date, parse_datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ParkingLot, ParkingTransaction, ParkingHistory, HourlyOccupancy
from .serializers import (SummarySerializer, RevenueLineSerializer, RevenueBarSerializer, ParkingLotSerializer,
                          ParkingLotDetailSerializer)
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter


class ParkingLotListView(APIView):
    def get(self, request):
        parking_lots = ParkingLot.objects.all()
        serializer = ParkingLotSerializer(parking_lots, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class HistoricalOccupancyView(APIView):
    def get(self, request, pk):
        selected_month = request.query_params.get('month')
        date = datetime.datetime.strptime(selected_month, "%Y-%m")
        occupancy_data = ParkingHistory.objects.filter(
            parking_lot_id=pk,
            date__year=date.year,
            date__month=date.month,
        ).values('date', 'occupancy_rate')

        return Response(occupancy_data, status=status.HTTP_200_OK)


class PeakHoursView(APIView):
    def get(self, request, pk):
        selected_date = request.query_params.get('date')
        peak_hours_data = HourlyOccupancy.objects.filter(
            parking_lot_id=pk,
            date=parse_date(selected_date) if selected_date else None
        ).values('hour', 'occupancy_rate')

        return Response(peak_hours_data, status=status.HTTP_200_OK)


class RevenueView(APIView):
    def get(self, request, pk):
        selected_month = request.query_params.get('month')
        date = datetime.datetime.strptime(selected_month, "%Y-%m")
        revenue_data = ParkingHistory.objects.filter(
            parking_lot_id=pk,
            date__year=date.year,
            date__month=date.month,
        ).values('date', 'total_revenue')

        return Response(revenue_data, status=status.HTTP_200_OK)


class MonthlyRevenueView(APIView):
    def get(self, request, pk):
        selected_year = request.query_params.get('year')
        monthly_revenue_data = ParkingHistory.objects.filter(
            parking_lot_id=pk,
            date__year=selected_year if selected_year else None
        ).annotate(month=TruncMonth('date')).values('month').annotate(total_revenue=Sum('total_revenue')).values('month', 'total_revenue')

        return Response(monthly_revenue_data, status=status.HTTP_200_OK)



class SummaryView(APIView):
    def get(self, request):
        total_revenue = ParkingHistory.objects.aggregate(totalRevenue=Sum('total_revenue'))['totalRevenue'] or 0
        total_lots = ParkingLot.objects.count()
        total_capacity = ParkingLot.objects.aggregate(totalCapacity=Sum('capacity'))['totalCapacity'] or 0

        summary_data = {
            'totalRevenue': total_revenue,
            'totalLots': total_lots,
            'totalCapacity': total_capacity,
        }
        serializer = SummarySerializer(summary_data)
        return Response(serializer.data)


class RevenueLineView(APIView):
    def get(self, request):
        try:
            # Aggregate monthly revenue data
            revenue_data = (
                ParkingHistory.objects
                .values('date__month', 'date__year')
                .annotate(monthly_revenue=Sum('total_revenue'))
                .order_by('date__year', 'date__month')
            )

            # Prepare the data for the response
            response_data = []
            for entry in revenue_data:
                month_year = datetime.date(entry['date__year'], entry['date__month'], 1).strftime('%B %Y')
                response_data.append({
                    'month': month_year,
                    'revenue': entry['monthly_revenue']
                })

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class RevenueBarView(APIView):
    def get(self, request):
        try:
            # Get the month and year from the request query parameters
            month = request.query_params.get('month')
            year = request.query_params.get('year')

            if not month or not year:
                return Response({'error': 'Month and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

            # Filter the ParkingHistory records for the selected month and year
            revenue_data = (
                ParkingHistory.objects
                .filter(date__year=year, date__month=month)
                .values('parking_lot__name')
                .annotate(monthly_revenue=Sum('total_revenue'))
                .order_by('parking_lot__name')
            )

            # Prepare the data for the response
            response_data = []
            for entry in revenue_data:
                response_data.append({
                    'lotName': entry['parking_lot__name'],
                    'revenue': entry['monthly_revenue']
                })

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class BatchImportParkingHistoryView(APIView):
    def post(self, request):
        # Check if the 'file' is in request.FILES
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        file = request.FILES['file']
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file.read()))
            sheet = workbook.active

            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                parking_lot_name, date_v, occupancy_rate, total_revenue = row

                try:
                    parking_lot = ParkingLot.objects.get(name=parking_lot_name)

                    # Check if date is already a datetime object
                    if isinstance(date_v, datetime.datetime):
                        date_v = date_v.date()  # Convert to date if it's a datetime
                    elif isinstance(date_v, str):
                        date_v = datetime.datetime.strptime(date_v, '%Y-%m-%d').date()

                    occupancy_rate = float(occupancy_rate) if occupancy_rate else None
                    total_revenue = float(total_revenue)

                    records.append(ParkingHistory(
                        parking_lot=parking_lot,
                        date=date_v,
                        occupancy_rate=occupancy_rate,
                        total_revenue=total_revenue
                    ))
                except ParkingLot.DoesNotExist:
                    return Response({'error': f"Parking lot {parking_lot_name} does not exist."}, status=status.HTTP_400_BAD_REQUEST)
                except Exception as e:
                    return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

            ParkingHistory.objects.bulk_create(records)
            return Response({'success': 'Data imported successfully!'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': f"Error processing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

def generate_parking_history_excel_template(request):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parking History Template"

    # Define the headers that match your ParkingHistory model
    headers = ["parking_lot", "date", "occupancy_rate", "total_revenue"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Set the content type to 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=parking_history_template.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


class BatchImportParkingTransactionView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)
            transactions = []
            for _, row in df.iterrows():
                lot_name = row['Parking Lot']
                lot = ParkingLot.objects.get(name=lot_name)
                transactions.append(ParkingTransaction(
                    parking_lot=lot,
                    license_plate=row['License Plate'],
                    entry_time=parse_datetime(row['Entry Time']),
                    exit_time=parse_datetime(row['Exit Time']),
                    revenue=row['Revenue']
                ))
            ParkingTransaction.objects.bulk_create(transactions)
            return Response({"status": "success"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


def generate_parking_transaction_excel_template(request):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parking Transaction Template"

    # Define the headers that match your ParkingTransaction model
    headers = ["Parking Lot", "License Plate", "Entry Time", "Exit Time", "Revenue"]

    # Add headers to the first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Set the content type to 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=parking_transaction_template.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response